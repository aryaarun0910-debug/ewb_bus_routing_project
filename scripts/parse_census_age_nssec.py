"""
parse_census_age_nssec.py
=========================
Parses Census 2021 TS007 (age by single year) and TS062 (NS-SeC / 16+ pop)
for Ladywood stops.

TS007 — downloaded at ward level (Ladywood ward col present)
  Extracts: under-16 %, working-age (16-64) %, elderly (65+) %
  These are area-wide ward figures applied to all stops in Ladywood.

TS062 — downloaded at LSOA level (total aged 16+ only)
  Extracts: working-age population count per LSOA.

Output
------
  data/census/ladywood_age_profile.json
  data/census/ladywood_working_age_pop.json

Usage
-----
  python scripts/parse_census_age_nssec.py
"""

from __future__ import annotations

import json
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

_REPO = Path(__file__).parent.parent

TS007_XLSX = _REPO / "data" / "census" / "ts007_age_birmingham_lsoa_2021.xlsx"
TS062_XLSX = _REPO / "data" / "census" / "ts062_travel_all_trips_birmingham_lsoa_2021.xlsx"
OUT_AGE    = _REPO / "data" / "census" / "ladywood_age_profile.json"
OUT_WAP    = _REPO / "data" / "census" / "ladywood_working_age_pop.json"

# LSOA21 codes, re-derived from corrected GTFS coords (see
# scripts/fetch_lsoa_population.py). S14's LSOA (E01010062) is in Sandwell,
# outside the Birmingham LA xlsx download scope.
STOP_LSOA: dict[str, dict] = {
    "S01": {"lsoa": "E01033615", "name": "New Street Station"},
    "S02": {"lsoa": "E01033624", "name": "Spring St"},
    "S03": {"lsoa": "E01033559", "name": "Jewellery Quarter Station"},
    "S04": {"lsoa": "E01033638", "name": "Soho Hill"},
    "S05": {"lsoa": "E01034948", "name": "Five Ways (Metro)"},
    "S06": {"lsoa": "E01009153", "name": "Dudley Rd"},
    "S07": {"lsoa": "E01033626", "name": "Five Ways Station"},
    "S08": {"lsoa": "E01009143", "name": "Icknield Port Rd"},
    "S09": {"lsoa": "E01033640", "name": "Belgrave Interchange"},
    "S10": {"lsoa": "E01009140", "name": "Ladywood Fire Station"},
    "S11": {"lsoa": "E01009143", "name": "Edgbaston Village Metro"},
    "S12": {"lsoa": "E01009152", "name": "Summerfield Park"},
    "S13": {"lsoa": "E01009346", "name": "City Rd Medical Centre"},
    "S14": {"lsoa": "E01010062", "name": "Mencap Centre"},
    "S15": {"lsoa": "E01009153", "name": "Summerfield Crescent"},
}


def parse_ts007() -> dict:
    """Extract Ladywood ward age profile from TS007."""
    wb = openpyxl.load_workbook(TS007_XLSX)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Find header row and Ladywood column
    header_idx = None
    ladywood_col = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell and "ladywood" in str(cell).lower():
                header_idx = i
                ladywood_col = j
                break
        if header_idx is not None:
            break

    if ladywood_col is None:
        print("TS007: Ladywood column not found")
        return {}

    print(f"TS007: Ladywood ward at col {ladywood_col}, header row {header_idx}")

    # Parse age rows
    age_data: dict[str, int] = {}
    for row in rows[header_idx + 1:]:
        label = str(row[0]).strip() if row[0] else ""
        val   = row[ladywood_col]
        if not label or val is None:
            continue
        try:
            age_data[label] = int(val)
        except (TypeError, ValueError):
            pass

    total = age_data.get("Total: All usual residents", 0) or 1

    # Age bands
    under_16  = sum(v for k, v in age_data.items()
                    if any(f"aged {n}" in k.lower() for n in
                           ["under 1", "1 year", "2 year", "3 year", "4 year",
                            "5 year", "6 year", "7 year", "8 year", "9 year",
                            "10 year", "11 year", "12 year", "13 year", "14 year", "15 year"]))
    over_64   = sum(v for k, v in age_data.items()
                    if any(f"aged {n}" in k.lower() for n in
                           [str(x) for x in range(65, 101)]))
    working   = total - under_16 - over_64

    result = {
        "geography":        "Ladywood ward (ward2022)",
        "data_source":      "Census 2021 TS007 (Nomis)",
        "total_residents":  total,
        "under_16":         under_16,
        "working_age_16_64": working,
        "elderly_65_plus":  over_64,
        "under_16_pct":     round(100 * under_16 / total, 1),
        "working_age_pct":  round(100 * working  / total, 1),
        "elderly_pct":      round(100 * over_64  / total, 1),
        "note": "Ward-level figures — applied to all stops in Ladywood ward",
    }

    print(f"\nLadywood Ward Age Profile:")
    print(f"  Total residents : {total:,}")
    print(f"  Under 16        : {under_16:,} ({result['under_16_pct']}%)")
    print(f"  Working age     : {working:,} ({result['working_age_pct']}%)")
    print(f"  65+             : {over_64:,} ({result['elderly_pct']}%)")
    return result


def parse_ts062() -> dict:
    """Extract 16+ population per LSOA from TS062."""
    wb = openpyxl.load_workbook(TS062_XLSX)
    ws = wb.active

    lsoa_pop: dict[str, int] = {}
    for row in ws.iter_rows(values_only=True):
        if not row[0] or not isinstance(row[0], str):
            continue
        code = row[0].split(" : ")[0].strip()
        if not code.startswith("E01"):
            continue
        try:
            lsoa_pop[code] = int(row[1]) if row[1] else 0
        except (TypeError, ValueError):
            pass

    print(f"\nTS062: loaded {len(lsoa_pop)} Birmingham LSOAs (aged 16+)")

    result = {}
    matched = 0
    for sid, info in STOP_LSOA.items():
        lsoa = info["lsoa"]
        pop16 = lsoa_pop.get(lsoa)
        if pop16 is None:
            result[sid] = {"lsoa": lsoa, "stop_name": info["name"],
                           "note": "not in Birmingham LA download"}
            continue
        result[sid] = {
            "lsoa":           lsoa,
            "stop_name":      info["name"],
            "pop_aged_16plus": pop16,
            "data_source":    "Census 2021 TS062 (Nomis)",
        }
        matched += 1

    sep = "-" * 55
    print(f"\n{'Working-Age Population (16+) by Stop':^55}")
    print(sep)
    for sid in sorted(result, key=lambda s: result[s].get("pop_aged_16plus", 0), reverse=True):
        r = result[sid]
        if "pop_aged_16plus" not in r:
            print(f"  {sid}  {r['stop_name']:<30}  no data")
        else:
            bar = "#" * (r["pop_aged_16plus"] // 50)
            print(f"  {sid}  {r['stop_name']:<30}  {r['pop_aged_16plus']:>5}  {bar}")
    print(sep)
    print(f"Matched {matched}/15 stops")
    return result


if __name__ == "__main__":
    age_profile = parse_ts007()
    wap_data    = parse_ts062()

    OUT_AGE.write_text(json.dumps(age_profile, indent=2), encoding="utf-8")
    OUT_WAP.write_text(json.dumps(wap_data,    indent=2), encoding="utf-8")
    print(f"\nWrote {OUT_AGE}")
    print(f"Wrote {OUT_WAP}")
